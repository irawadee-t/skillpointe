"""Generate the SKILLED Nation career taxonomy artifacts from Tasha's workbook.

Single source of truth: "Industry & Career List Revisions v2" (xlsx). This
script reads it, applies the documented cleanups, and emits three artifacts
that must never drift from each other:

  1. packages/matching/sn_taxonomy.py   — pure-python data + derived adjacency
     (the matching engine imports this; no DB I/O, fully unit-testable)
  2. supabase/migrations/<ts>_skilled_nation_taxonomy.sql — sectors table,
     career-field upserts into canonical_job_families, validation trigger
  3. apps/web/src/lib/taxonomy.generated.ts — sector/field constants for the
     cascading selects (no fetch needed at form time)

Documented cleanups (each verified against the workbook):
  * "Welding & Fabricating" (Construction/Transportation tabs) is the same
    field as "Welding and Fabrication" (matrix + Manufacturing tab). The
    matrix name wins; the tab spelling is kept as an alias.
  * "Sheet Metal Fabrication" appears twice in the matrix (a trailing-space
    duplicate row splits its memberships). Merged: Construction +
    Manufacturing + Transportation.
  * "Energy Storage" exists in the Energy & Utilities tab but not the matrix.
    Included (Energy & Utilities).
  * Source typos fixed in display names, retained as aliases:
    "Cardovascular Sonography" -> "Cardiovascular Sonography",
    "Mechancial Design, CAD/CAM & Drafting" -> "Mechanical Design, CAD/CAM & Drafting".
  * Every sector gets its "Other X-related field" escape option (is_other),
    exactly as the tabs specify. These carry sector-level matching signal only.

Usage:
  python scripts/gen_taxonomy.py --xlsx "<path to workbook>" [--migration-ts YYYYMMDDHHMMSS]
"""
from __future__ import annotations

import argparse
import json
import re
import unicodedata
from pathlib import Path

REPO = Path(__file__).resolve().parent.parent

# ---------------------------------------------------------------------------
# Sector canon: stable code, short display name, matrix header spelling(s)
# ---------------------------------------------------------------------------
SECTORS = [
    # (code, short display name, matrix/tab surface forms)
    ("construction",  "Construction & Building Trades",
     {"construction & building trades"}),
    ("manufacturing", "Manufacturing", {"manufacturing"}),
    ("energy",        "Energy & Utilities", {"energy & utilities"}),
    ("transportation",
     "Transportation",
     {"transportation (automotive, diesel, aviation, marine, rail, and related logistics)"}),
    ("healthcare",    "Healthcare", {"healthcare"}),
    ("data_it",       "Data & Information Technology",
     {"data & it", "data & information technology"}),
    ("telecom",       "Telecommunications", {"telecommunications"}),
    ("public_safety", "Public & Emergency Service",
     {"public & emergency service"}),
]
SECTOR_BY_SURFACE = {s: code for code, _n, surfaces in SECTORS for s in surfaces}

# The full sector name used on the Transportation card (kept for description).
TRANSPORT_FULL = ("Transportation (Automotive, Diesel, Aviation, Marine, Rail, "
                  "and Related Logistics)")

# Display-name corrections: source spelling -> canonical spelling.
NAME_FIXES = {
    "Cardovascular Sonography": "Cardiovascular Sonography",
    "Mechancial Design, CAD/CAM & Drafting": "Mechanical Design, CAD/CAM & Drafting",
}
# Fields whose auto-slug is unwieldy; explicit stable codes.
CODE_OVERRIDES = {
    "Aircraft/Aviation Maintenance (including A&P Mechanic)": "aviation_maintenance",
    "Automotive and EV Service & Technology": "automotive_ev_service",
    "CNC Machining & Precision Machining": "cnc_machining",
    "Electrical (Residential or Commercial)": "electrical",
    "Electrical/Transmission Linework": "transmission_linework",
    "Geographic Information Systems (GIS)": "gis",
    "Industrial Maintenance, Millwright & Mechanical Systems": "industrial_maintenance",
    "Inspection, Metrology & CMM Operation": "metrology_cmm",
    "Instrumentation, Automation, & Controls": "instrumentation_automation_controls",
    "Instrumentation and Controls": "instrumentation_controls",
    "Interior Finishing (e.g., Flooring, Drywall, Insulation, Painting, Interior Design)":
        "interior_finishing",
    "Manufacturing Engineering Technology & Industrial Technology": "manufacturing_engineering_tech",
    "Manufacturing Production & Machine Operation": "manufacturing_production",
    "Mechanical Design, CAD/CAM & Drafting": "mechanical_design_cad_cam",
    "Medical Records and Health Information": "health_information",
    "Motorcycle / Powersports Service & Technology": "powersports_service",
    "Nursing (LPN/LVN, ADN, RN)": "nursing",
    "Nursing Assistant (CNA)": "nursing_assistant",
    "Oil / Natural Gas Production Technology": "oil_gas_production",
    "Power Plant Technology/Operation": "power_plant_operation",
    "Rail Signals, Communications, and Controls": "rail_signals_controls",
    "Railway Track Construction and Maintenance": "railway_track_maintenance",
    "Security System Technology / Locksmithing": "security_systems_locksmithing",
    "Tool, Die, Mold & Fixture Making": "tool_die_mold",
    "Plastics, Polymers & Composites Manufacturing": "plastics_composites",
    "Materials Testing & Nondestructive Testing": "materials_ndt",
    "Additive Manufacturing & 3D Printing": "additive_manufacturing",
    "Commercial Driving (CDL)": "commercial_driving_cdl",
    "Sustainable/Renewable Energy": "renewable_energy",
    "Solar Installation and Maintenance": "solar_installation",
    "Pipeline Construction and Operation": "pipeline_construction",
    "Welding and Fabrication": "welding_fabrication",
    "HVAC/R": "hvac_r",
    "Cosmetology / Esthetician": "cosmetology_esthetician",
    "Auto Body / Collision Repair": "auto_body_collision",
    "EMT / Paramedic": "emt_paramedic",
    "Law Enforcement / Criminology": "law_enforcement",
    "Utility/Public Works": "utility_public_works",
    "Waste/Wastewater Operations": "wastewater_operations",
    "Cable/Fiber Optics Technician": "fiber_optics_technician",
    "Robotics & Mechatronics": "robotics_mechatronics",
    "Diet & Nutrition": "diet_nutrition",
    "Surveying and Mapping": "surveying_mapping",
    "Marine Systems Service & Technology": "marine_systems_service",
    "Building Automation & Controls": "building_automation_controls",
    "IT & Network Support": "it_network_support",
}
# Per-sector "Other" escape options — exact labels from Tasha's tabs.
OTHER_LABELS = {
    "construction":  "Other Construction-related field",
    "energy":        "Other Energy/Utilities-related field",
    "manufacturing": "Other Manufacturing-related field",
    "transportation": "Other Transportation-related field",
    "healthcare":    "Other Healthcare-related field",
    "data_it":       "Other Data/IT-related field",
    "telecom":       "Other Telecommunications-related field",
    "public_safety": "Other Public & Emergency Services-related field",
}

# Legacy family-code bridge: pre-Tasha codes -> new field codes. Used by the
# remap script as the FALLBACK when the applicant/job raw text does not
# alias-match a new field directly. `None` = no honest home; route to review.
LEGACY_BRIDGE = {
    "electrical": "electrical", "plumbing": "plumbing", "hvac": "hvac_r",
    "construction": "building_construction_technology",
    "welding": "welding_fabrication", "automotive": "automotive_ev_service",
    "manufacturing": "manufacturing_production",
    "logistics": "other_transportation",
    "heavy_equipment": "heavy_equipment_operation",
    "security": "security_systems_locksmithing",
    "drafting": "architectural_drafting_cad",
    "aviation": "aviation_maintenance", "auto_body": "auto_body_collision",
    "aviation_aerospace": "aviation_maintenance",
    "energy_lineman": "transmission_linework",
    "solar_energy": "solar_installation", "wind_energy": "wind_turbine_technology",
    "robotics": "robotics_mechatronics", "construction_mgmt": "construction_management",
    "healthcare_support": "patient_care", "dental": "dental_assistant",
    "nursing": "nursing", "radiology": "radiology_technician",
    "respiratory": "respiratory_therapy", "physical_therapy": "physical_therapy_assistant",
    "pharmacy": "pharmacy_technician", "surgical_tech": "surgical_technology",
    "veterinary": "veterinary_technician", "lab_sciences": "laboratory_technician",
    "health_information": "health_information", "dietetics": "diet_nutrition",
    "civil_survey": "surveying_mapping", "field_service": "industrial_maintenance",
    "rail_transit": "rail_vehicle_maintenance", "marine": "marine_systems_service",
    "power_plant": "power_plant_operation",
    "building_automation": "building_automation_controls",
    "data_center": "data_center_operations",
    "industrial_maintenance": "industrial_maintenance",
    "electronics": "industrial_electrical_technology",
    "it_support": "it_network_support",
    "administrative": None,  # no honest home in the trades taxonomy -> review
    # packages/scraper/trades.py classifier vocabulary (title-based):
    "utilities_energy": "utility_public_works",
    "logistics_warehouse": "other_transportation",
    "automotive_diesel": "diesel_service_and_technology",
    "machining_cnc": "cnc_machining",
    "construction_skilled": "building_construction_technology",
}


def norm(s: str) -> str:
    s = unicodedata.normalize("NFKC", s).strip()
    return re.sub(r"\s+", " ", s)


def slugify(name: str) -> str:
    if name in CODE_OVERRIDES:
        return CODE_OVERRIDES[name]
    s = name.lower()
    s = s.replace("&", " and ")
    s = re.sub(r"[/,()\.]", " ", s)
    s = re.sub(r"[^a-z0-9 ]", "", s)
    return "_".join(s.split())


def extract(xlsx: Path) -> dict:
    import openpyxl
    wb = openpyxl.load_workbook(xlsx, data_only=True)

    # Sector descriptions/examples (full text) from the Sectors sheet.
    meta = {}
    for row in wb["Sectors"].iter_rows(min_row=2, values_only=True):
        _cat, industry, desc, examples = (tuple(row) + (None,) * 4)[:4]
        if industry and str(industry).strip():
            code = SECTOR_BY_SURFACE[norm(str(industry)).lower()]
            meta[code] = {
                "description": norm(str(desc)) if desc else "",
                "examples": norm(str(examples)) if examples else "",
            }

    # Careers matrix -> field -> set of sector codes (merging duplicates).
    ws = wb["Careers (Full List)"]
    rows = list(ws.iter_rows(values_only=True))
    header = [norm(str(c)) if c else "" for c in rows[0]]
    col_sectors = [SECTOR_BY_SURFACE[h.lower()] for h in header[1:] if h]

    fields: dict[str, set[str]] = {}
    for row in rows[1:]:
        raw = row[0]
        if not raw or not str(raw).strip():
            continue
        name = norm(str(raw))
        name = NAME_FIXES.get(name, name)
        memberships = {
            col_sectors[i]
            for i, cell in enumerate(row[1:1 + len(col_sectors)])
            if cell and str(cell).strip().upper() == "Y"
        }
        fields.setdefault(name, set()).update(memberships)

    # Tab-only addition, verified against the workbook.
    fields.setdefault("Energy Storage", set()).add("energy")

    # Source-spelling aliases for ETL/normalization continuity.
    aliases = {
        "Welding and Fabrication": ["Welding & Fabricating"],
        "Cardiovascular Sonography": ["Cardovascular Sonography"],
        "Mechanical Design, CAD/CAM & Drafting": ["Mechancial Design, CAD/CAM & Drafting"],
    }
    return {"meta": meta, "fields": fields, "aliases": aliases}


def build(xlsx: Path) -> dict:
    data = extract(xlsx)
    out_fields = []
    for name in sorted(data["fields"]):
        sectors = sorted(data["fields"][name])
        assert sectors, f"field with no sector: {name!r}"
        out_fields.append({
            "code": slugify(name),
            "name": name,
            "sectors": sectors,
            "is_other": False,
            "aliases": data["aliases"].get(name, []),
        })
    for code, _n, _s in SECTORS:
        out_fields.append({
            "code": f"other_{code}",
            "name": OTHER_LABELS[code],
            "sectors": [code],
            "is_other": True,
            "aliases": [],
        })
    codes = [f["code"] for f in out_fields]
    assert len(codes) == len(set(codes)), "duplicate field codes"
    code_set = set(codes)
    bad_bridge = {
        old: new for old, new in LEGACY_BRIDGE.items()
        if new is not None and new not in code_set
    }
    assert not bad_bridge, f"LEGACY_BRIDGE targets missing from FIELDS: {bad_bridge}"
    return {
        "sectors": [
            {"code": c, "name": n,
             "full_name": TRANSPORT_FULL if c == "transportation" else n,
             **data["meta"].get(c, {"description": "", "examples": ""})}
            for c, n, _s in SECTORS
        ],
        "fields": out_fields,
    }


# ---------------------------------------------------------------------------
# Emitters
# ---------------------------------------------------------------------------

def emit_python(tax: dict, path: Path) -> None:
    lines = [
        '"""SKILLED Nation career taxonomy — GENERATED, do not edit by hand.',
        "",
        "Source: Industry & Career List Revisions v2 (Tasha). Regenerate with",
        "scripts/gen_taxonomy.py. Pure data + derived adjacency; no I/O.",
        '"""',
        "from __future__ import annotations",
        "",
        "SECTORS: dict[str, dict] = {",
    ]
    for s in tax["sectors"]:
        lines.append(f'    "{s["code"]}": {{"name": {s["name"]!r}, "full_name": {s["full_name"]!r},')
        lines.append(f'        "description": {s["description"]!r},')
        lines.append(f'        "examples": {s["examples"]!r}}},')
    lines.append("}")
    lines.append("")
    lines.append("FIELDS: dict[str, dict] = {")
    for f in tax["fields"]:
        lines.append(
            f'    "{f["code"]}": {{"name": {f["name"]!r}, "sectors": {f["sectors"]!r}, '
            f'"is_other": {f["is_other"]}, "aliases": {f["aliases"]!r}}},'
        )
    lines.append("}")
    lines.append("")
    lines.append('LEGACY_FAMILY_BRIDGE: dict[str, str | None] = {')
    for old, new in LEGACY_BRIDGE.items():
        lines.append(f'    "{old}": {new!r},')
    lines.append("}")
    lines.append("")
    lines.append('''

def resolve_field_code(code: str | None) -> str | None:
    """Map any known code (new or legacy) to a current field code."""
    if code is None:
        return None
    if code in FIELDS:
        return code
    return LEGACY_FAMILY_BRIDGE.get(code)


def field_sectors(code: str | None) -> set[str]:
    f = FIELDS.get(resolve_field_code(code) or "")
    return set(f["sectors"]) if f else set()


def fields_share_sector(a: str | None, b: str | None) -> bool:
    return bool(field_sectors(a) & field_sectors(b))


def relate(a: str | None, b: str | None) -> str:
    """Relationship between two field codes (legacy codes resolve first).

    "same"      — identical named field (an "Other X" pair never counts as
                  same: it carries sector-level signal only)
    "adjacent"  — distinct fields sharing at least one sector
    "unrelated" — no shared sector
    "unknown"   — either side missing or not in the taxonomy
    """
    ra, rb = resolve_field_code(a), resolve_field_code(b)
    if ra is None or rb is None:
        return "unknown"
    if ra == rb and not FIELDS[ra]["is_other"]:
        return "same"
    if set(FIELDS[ra]["sectors"]) & set(FIELDS[rb]["sectors"]):
        return "adjacent"
    return "unrelated"


# Derived adjacency: two distinct fields are adjacent iff they share a sector.
# "Other" fields are adjacent to everything in their sector by construction.
FIELD_ADJACENCY: dict[str, set[str]] = {}
for _code, _f in FIELDS.items():
    FIELD_ADJACENCY[_code] = {
        _c for _c, _g in FIELDS.items()
        if _c != _code and set(_g["sectors"]) & set(_f["sectors"])
    }
''')
    path.write_text("\n".join(lines))


def emit_sql(tax: dict, path: Path) -> None:
    def q(s: str) -> str:
        return s.replace("'", "''")

    parts = [
        "-- GENERATED by scripts/gen_taxonomy.py from 'Industry & Career List Revisions v2'.",
        "-- SKILLED Nation two-level taxonomy: sectors + career fields.",
        "",
        "CREATE TABLE IF NOT EXISTS public.sectors (",
        "  code          TEXT PRIMARY KEY,",
        "  name          TEXT NOT NULL,",
        "  full_name     TEXT NOT NULL,",
        "  description   TEXT NOT NULL DEFAULT '',",
        "  examples      TEXT NOT NULL DEFAULT '',",
        "  display_order INT  NOT NULL,",
        "  created_at    TIMESTAMPTZ NOT NULL DEFAULT NOW()",
        ");",
        "",
    ]
    for i, s in enumerate(tax["sectors"]):
        parts.append(
            "INSERT INTO public.sectors (code, name, full_name, description, examples, display_order)\n"
            f"VALUES ('{s['code']}', '{q(s['name'])}', '{q(s['full_name'])}', "
            f"'{q(s['description'])}', '{q(s['examples'])}', {i})\n"
            "ON CONFLICT (code) DO UPDATE SET name = EXCLUDED.name, full_name = EXCLUDED.full_name,\n"
            "  description = EXCLUDED.description, examples = EXCLUDED.examples,\n"
            "  display_order = EXCLUDED.display_order;"
        )
    parts += [
        "",
        "ALTER TABLE public.canonical_job_families ADD COLUMN IF NOT EXISTS is_other BOOLEAN NOT NULL DEFAULT FALSE;",
        "",
        "-- Career fields land in canonical_job_families (code = field code,",
        "-- industries[] = sector codes) so every existing FK keeps working.",
    ]
    for f in tax["fields"]:
        sect = "{" + ",".join(f["sectors"]) + "}"
        alias_sql = ("ARRAY[" + ",".join(f"'{q(a)}'" for a in [f["name"], *f["aliases"]]) + "]::text[]")
        parts.append(
            "INSERT INTO public.canonical_job_families (code, name, aliases, industries, is_other, is_active)\n"
            f"VALUES ('{f['code']}', '{q(f['name'])}', {alias_sql}, '{sect}', {str(f['is_other']).upper()}, TRUE)\n"
            "ON CONFLICT (code) DO UPDATE SET name = EXCLUDED.name,\n"
            "  industries = EXCLUDED.industries, is_other = EXCLUDED.is_other, is_active = TRUE,\n"
            "  aliases = (SELECT array_agg(DISTINCT a) FROM unnest(public.canonical_job_families.aliases || EXCLUDED.aliases) AS a);"
        )
    parts += [
        "",
        "-- Retire codes from superseded auto-slug runs (renamed via overrides).",
        "UPDATE public.canonical_job_families SET is_active = FALSE WHERE code IN (",
        "  'robotics_and_mechatronics', 'diet_and_nutrition', 'surveying_and_mapping',",
        "  'marine_systems_service_and_technology', 'building_automation_and_controls',",
        "  'it_and_network_support');",
        "",
        "-- Sector labels on the two entities that carry a field.",
        "ALTER TABLE public.applicants ADD COLUMN IF NOT EXISTS sector_code TEXT REFERENCES public.sectors(code);",
        "ALTER TABLE public.jobs       ADD COLUMN IF NOT EXISTS sector_code TEXT REFERENCES public.sectors(code);",
        "",
        "-- Integrity: a chosen sector must be one the linked field belongs to.",
        "CREATE OR REPLACE FUNCTION public.check_sector_matches_field() RETURNS trigger AS $$",
        "DECLARE fam_industries TEXT[];",
        "BEGIN",
        "  IF NEW.sector_code IS NULL OR NEW.canonical_job_family_id IS NULL THEN RETURN NEW; END IF;",
        "  SELECT industries INTO fam_industries FROM public.canonical_job_families WHERE id = NEW.canonical_job_family_id;",
        "  IF fam_industries IS NOT NULL AND array_length(fam_industries, 1) IS NOT NULL",
        "     AND NOT (NEW.sector_code = ANY(fam_industries)) THEN",
        "    RAISE EXCEPTION 'sector % is not valid for the selected field', NEW.sector_code",
        "      USING ERRCODE = '23514';",
        "  END IF;",
        "  RETURN NEW;",
        "END; $$ LANGUAGE plpgsql;",
        "",
        "DROP TRIGGER IF EXISTS applicants_sector_field_check ON public.applicants;",
        "CREATE TRIGGER applicants_sector_field_check BEFORE INSERT OR UPDATE OF sector_code, canonical_job_family_id",
        "  ON public.applicants FOR EACH ROW EXECUTE FUNCTION public.check_sector_matches_field();",
        "DROP TRIGGER IF EXISTS jobs_sector_field_check ON public.jobs;",
        "CREATE TRIGGER jobs_sector_field_check BEFORE INSERT OR UPDATE OF sector_code, canonical_job_family_id",
        "  ON public.jobs FOR EACH ROW EXECUTE FUNCTION public.check_sector_matches_field();",
        "",
    ]
    path.write_text("\n".join(parts) + "\n")


def emit_ts(tax: dict, path: Path) -> None:
    lines = [
        "/** SKILLED Nation career taxonomy — GENERATED by scripts/gen_taxonomy.py.",
        " *  Source: Industry & Career List Revisions v2 (Tasha). Do not edit. */",
        "",
        "export interface Sector { code: string; name: string; description: string; examples: string }",
        "export interface CareerField { code: string; name: string; sectors: string[]; isOther: boolean }",
        "",
        "export const SECTORS: Sector[] = [",
    ]
    for s in tax["sectors"]:
        lines.append(
            f'  {{ code: {json.dumps(s["code"])}, name: {json.dumps(s["name"])}, '
            f'description: {json.dumps(s["description"])}, examples: {json.dumps(s["examples"])} }},'
        )
    lines.append("];")
    lines.append("")
    lines.append("export const CAREER_FIELDS: CareerField[] = [")
    for f in tax["fields"]:
        lines.append(
            f'  {{ code: {json.dumps(f["code"])}, name: {json.dumps(f["name"])}, '
            f'sectors: {json.dumps(f["sectors"])}, isOther: {json.dumps(f["is_other"])} }},'
        )
    lines.append("];")
    lines.append("""
/** Fields selectable under a sector, "Other" pinned last, rest alphabetical. */
export function fieldsForSector(sectorCode: string): CareerField[] {
  const inSector = CAREER_FIELDS.filter((f) => f.sectors.includes(sectorCode));
  const named = inSector.filter((f) => !f.isOther).sort((a, b) => a.name.localeCompare(b.name));
  const other = inSector.filter((f) => f.isOther);
  return [...named, ...other];
}

export function fieldBelongsToSector(fieldCode: string, sectorCode: string): boolean {
  const f = CAREER_FIELDS.find((x) => x.code === fieldCode);
  return !!f && f.sectors.includes(sectorCode);
}
""")
    path.write_text("\n".join(lines))


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--xlsx", required=True)
    ap.add_argument("--migration-ts", default=None)
    args = ap.parse_args()

    tax = build(Path(args.xlsx))
    n_named = sum(1 for f in tax["fields"] if not f["is_other"])
    n_other = sum(1 for f in tax["fields"] if f["is_other"])
    multi = [f for f in tax["fields"] if len(f["sectors"]) > 1]
    print(f"sectors={len(tax['sectors'])} named_fields={n_named} other_fields={n_other} multi_sector={len(multi)}")

    py_path = REPO / "packages/matching/sn_taxonomy.py"
    emit_python(tax, py_path)
    print("wrote", py_path)

    if args.migration_ts:
        sql_path = REPO / f"supabase/migrations/{args.migration_ts}_skilled_nation_taxonomy.sql"
        emit_sql(tax, sql_path)
        print("wrote", sql_path)

    ts_path = REPO / "apps/web/src/lib/taxonomy.generated.ts"
    emit_ts(tax, ts_path)
    print("wrote", ts_path)

    # Provenance snapshot for review.
    json.dump(tax, open(REPO / "docs/taxonomy_snapshot.json", "w"), indent=1)
    print("wrote docs/taxonomy_snapshot.json")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
